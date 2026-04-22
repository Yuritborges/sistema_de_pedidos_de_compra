# app/ui/widgets/relatorio_obra_excel.py
# Gera relatório Excel por obra.
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Cores ─────────────────────────────────────────────────────────────────────
COR_VERMELHO   = "C0392B"
COR_CINZA_ESC  = "2C2C2C"
COR_CINZA_MED  = "666666"
COR_CINZA_CLR  = "F0EDED"
COR_BRANCO     = "FFFFFF"
COR_VERDE      = "1E8449"
COR_AZUL       = "2980B9"
COR_ALT1       = "FBF7F7"
COR_ALT2       = "FFFFFF"
COR_HDR_TAB    = "1A1A1A"
COR_AMARELO    = "FFF3CD"


def _fill(hex_cor):
    return PatternFill("solid", fgColor=hex_cor)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def _border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom():
    s = Side(style="thin", color="CCCCCC")
    return Border(bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _fmt_val(v):
    try:
        return round(float(v or 0), 2)
    except Exception:
        return 0.0


def gerar_excel(caminho, nome_obra, dados_obra, pedidos, itens_por_pedido):
    """
    Gera arquivo .xlsx com 3 abas:
        1. Resumo    — dados da obra + cards + top fornecedores + resumo por empresa
        2. Pedidos   — tabela de todos os pedidos (uma linha por pedido)
        3. Itens     — tabela de todos os itens (uma linha por item, com referência ao pedido)
    """
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # ABA 1 — RESUMO
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 20

    row = 1

    # Título principal
    ws1.merge_cells(f"A{row}:E{row}")
    ws1[f"A{row}"] = "RELATÓRIO DE AUDITORIA DE OBRA"
    ws1[f"A{row}"].font      = _font(bold=True, size=16, color=COR_BRANCO)
    ws1[f"A{row}"].fill      = _fill(COR_VERMELHO)
    ws1[f"A{row}"].alignment = _align("center")
    ws1.row_dimensions[row].height = 28
    row += 1

    # Subtítulo obra
    ws1.merge_cells(f"A{row}:E{row}")
    ws1[f"A{row}"] = nome_obra.upper()
    ws1[f"A{row}"].font      = _font(bold=True, size=13, color=COR_CINZA_ESC)
    ws1[f"A{row}"].fill      = _fill(COR_CINZA_CLR)
    ws1[f"A{row}"].alignment = _align("center")
    ws1.row_dimensions[row].height = 22
    row += 1

    # Data geração
    ws1.merge_cells(f"A{row}:E{row}")
    ws1[f"A{row}"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    ws1[f"A{row}"].font      = _font(size=9, color=COR_CINZA_MED, italic=True)
    ws1[f"A{row}"].alignment = _align("right")
    row += 2

    # ── Dados da obra ─────────────────────────────────────────────────────────
    def add_par(label, valor):
        nonlocal row
        ws1[f"A{row}"] = label
        ws1[f"A{row}"].font = _font(bold=True, size=10)
        ws1.merge_cells(f"B{row}:E{row}")
        ws1[f"B{row}"] = valor or "—"
        ws1[f"B{row}"].font = _font(size=10)
        ws1.row_dimensions[row].height = 16
        row += 1

    ws1.merge_cells(f"A{row}:E{row}")
    ws1[f"A{row}"] = "DADOS DA OBRA"
    ws1[f"A{row}"].font      = _font(bold=True, size=10, color=COR_BRANCO)
    ws1[f"A{row}"].fill      = _fill(COR_VERMELHO)
    ws1[f"A{row}"].alignment = _align("left")
    ws1.row_dimensions[row].height = 18
    row += 1

    add_par("Escola / Descrição:", dados_obra.get("escola",""))
    add_par("Endereço:",           dados_obra.get("endereco",""))
    add_par("Bairro:",             dados_obra.get("bairro",""))
    add_par("Cidade / UF:",        f"{dados_obra.get('cidade','')} — {dados_obra.get('uf','SP')}")
    add_par("CEP:",                dados_obra.get("cep",""))
    add_par("Faturamento:",        dados_obra.get("faturamento",""))
    add_par("Contrato:",           dados_obra.get("contrato",""))
    row += 1

    # ── Cards de resumo ───────────────────────────────────────────────────────
    total_gasto  = sum(_fmt_val(p["valor_total"]) for p in pedidos)
    n_pedidos    = len(pedidos)
    total_itens  = sum(len(v) for v in itens_por_pedido.values())
    fornecedores = len({str(p["fornecedor_nome"] or "") for p in pedidos if p["fornecedor_nome"]})

    cards = [
        ("PEDIDOS EMITIDOS",   str(n_pedidos),                   COR_VERMELHO),
        ("VALOR TOTAL GASTO",  f"R$ {_fmt_brl(total_gasto)}",    COR_VERDE),
        ("TOTAL DE ITENS",     str(total_itens),                  COR_AZUL),
        ("FORNECEDORES",       str(fornecedores),                 "784212"),
    ]
    cols_cards = ["A","B","C","D"]
    for i, (lbl, val, cor) in enumerate(cards):
        c = cols_cards[i]
        ws1[f"{c}{row}"] = lbl
        ws1[f"{c}{row}"].font      = _font(bold=True, size=8, color=cor)
        ws1[f"{c}{row}"].fill      = _fill(COR_CINZA_CLR)
        ws1[f"{c}{row}"].alignment = _align("center")
        ws1.row_dimensions[row].height = 14

        ws1[f"{c}{row+1}"] = val
        ws1[f"{c}{row+1}"].font      = _font(bold=True, size=14, color=cor)
        ws1[f"{c}{row+1}"].fill      = _fill(COR_BRANCO)
        ws1[f"{c}{row+1}"].alignment = _align("center")
        ws1.row_dimensions[row+1].height = 22
    row += 3

    # ── Resumo por empresa ────────────────────────────────────────────────────
    from collections import defaultdict
    por_empresa = defaultdict(lambda: {"n":0,"v":0.0})
    for p in pedidos:
        try: emp = str(p["empresa_faturadora"] or "—")
        except: emp = "—"
        por_empresa[emp]["n"] += 1
        por_empresa[emp]["v"] += _fmt_val(p["valor_total"])

    ws1.merge_cells(f"A{row}:E{row}")
    ws1[f"A{row}"] = "RESUMO POR EMPRESA FATURADORA"
    ws1[f"A{row}"].font = _font(bold=True, size=10, color=COR_BRANCO)
    ws1[f"A{row}"].fill = _fill(COR_CINZA_ESC)
    ws1[f"A{row}"].alignment = _align("left")
    ws1.row_dimensions[row].height = 18
    row += 1

    for hdr, col in [("Empresa","A"),("Pedidos","B"),("Valor Total","C")]:
        ws1[f"{col}{row}"] = hdr
        ws1[f"{col}{row}"].font = _font(bold=True, size=9, color=COR_BRANCO)
        ws1[f"{col}{row}"].fill = _fill("444444")
        ws1[f"{col}{row}"].alignment = _align("center")
    ws1.row_dimensions[row].height = 16
    row += 1

    for i, (emp, info) in enumerate(sorted(por_empresa.items(), key=lambda x:-x[1]["v"])):
        bg = COR_ALT1 if i%2==0 else COR_ALT2
        ws1[f"A{row}"] = emp
        ws1[f"A{row}"].font = _font(bold=True, size=10)
        ws1[f"A{row}"].fill = _fill(bg)
        ws1[f"B{row}"] = info["n"]
        ws1[f"B{row}"].font = _font(size=10)
        ws1[f"B{row}"].fill = _fill(bg)
        ws1[f"B{row}"].alignment = _align("center")
        ws1[f"C{row}"] = info["v"]
        ws1[f"C{row}"].font = _font(bold=True, size=10, color=COR_VERDE)
        ws1[f"C{row}"].fill = _fill(bg)
        ws1[f"C{row}"].number_format = 'R$ #,##0.00'
        ws1[f"C{row}"].alignment = _align("right")
        ws1.row_dimensions[row].height = 16
        row += 1
    row += 1

    # ── Top fornecedores ──────────────────────────────────────────────────────
    por_forn = defaultdict(lambda: {"n":0,"v":0.0})
    for p in pedidos:
        try: fn = str(p["fornecedor_nome"] or "—").upper()
        except: fn = "—"
        por_forn[fn]["n"] += 1
        por_forn[fn]["v"] += _fmt_val(p["valor_total"])

    ws1.merge_cells(f"A{row}:E{row}")
    ws1[f"A{row}"] = "TOP FORNECEDORES POR VALOR GASTO"
    ws1[f"A{row}"].font = _font(bold=True, size=10, color=COR_BRANCO)
    ws1[f"A{row}"].fill = _fill(COR_CINZA_ESC)
    ws1[f"A{row}"].alignment = _align("left")
    ws1.row_dimensions[row].height = 18
    row += 1

    for hdr, col in [("Fornecedor","A"),("Pedidos","B"),("Valor Total","C"),("% do Total","D")]:
        ws1[f"{col}{row}"] = hdr
        ws1[f"{col}{row}"].font = _font(bold=True, size=9, color=COR_BRANCO)
        ws1[f"{col}{row}"].fill = _fill("444444")
        ws1[f"{col}{row}"].alignment = _align("center")
    ws1.row_dimensions[row].height = 16
    row += 1

    for i, (fn, info) in enumerate(sorted(por_forn.items(), key=lambda x:-x[1]["v"])[:20]):
        bg = COR_ALT1 if i%2==0 else COR_ALT2
        pct = (info["v"]/total_gasto*100) if total_gasto > 0 else 0
        ws1[f"A{row}"] = fn
        ws1[f"A{row}"].font = _font(size=10)
        ws1[f"A{row}"].fill = _fill(bg)
        ws1[f"B{row}"] = info["n"]
        ws1[f"B{row}"].font = _font(size=10)
        ws1[f"B{row}"].fill = _fill(bg)
        ws1[f"B{row}"].alignment = _align("center")
        ws1[f"C{row}"] = info["v"]
        ws1[f"C{row}"].font = _font(bold=True, size=10, color=COR_VERDE)
        ws1[f"C{row}"].fill = _fill(bg)
        ws1[f"C{row}"].number_format = 'R$ #,##0.00'
        ws1[f"C{row}"].alignment = _align("right")
        ws1[f"D{row}"] = round(pct/100, 4)
        ws1[f"D{row}"].font = _font(size=10, color=COR_CINZA_MED)
        ws1[f"D{row}"].fill = _fill(bg)
        ws1[f"D{row}"].number_format = '0.0%'
        ws1[f"D{row}"].alignment = _align("center")
        ws1.row_dimensions[row].height = 16
        row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # ABA 2 — PEDIDOS
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Pedidos")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"

    colunas_ped = [
        ("Nº Pedido",    12), ("Data",       12), ("Fornecedor",  30),
        ("Empresa",      14), ("Cond. Pgto", 14), ("Forma Pgto",  12),
        ("Prazo (dias)", 12), ("Valor Total", 16), ("Itens",       8),
    ]
    for i, (hdr, larg) in enumerate(colunas_ped, 1):
        col = get_column_letter(i)
        ws2.column_dimensions[col].width = larg
        ws2[f"{col}1"] = hdr
        ws2[f"{col}1"].font      = _font(bold=True, size=9, color=COR_BRANCO)
        ws2[f"{col}1"].fill      = _fill(COR_HDR_TAB)
        ws2[f"{col}1"].alignment = _align("center")
        ws2[f"{col}1"].border    = _border_thin()
    ws2.row_dimensions[1].height = 20

    for i, p in enumerate(pedidos, 2):
        bg = COR_ALT1 if i%2==0 else COR_ALT2
        def _pg(key, default=""):
            try: v = p[key]; return v if v is not None else default
            except: return default

        num = str(_pg("numero",""))
        n_itens_ped = len(itens_por_pedido.get(num, []))

        vals = [
            f"#{num}", _pg("data_pedido"), _pg("fornecedor_nome"),
            _pg("empresa_faturadora"), _pg("condicao_pagamento"),
            _pg("forma_pagamento"),    _pg("prazo_entrega",0),
            _fmt_val(_pg("valor_total",0)), n_itens_ped,
        ]
        for j, val in enumerate(vals, 1):
            col = get_column_letter(j)
            ws2[f"{col}{i}"] = val
            ws2[f"{col}{i}"].fill   = _fill(bg)
            ws2[f"{col}{i}"].border = _border_bottom()
            ws2[f"{col}{i}"].font   = _font(size=10)
            if j == 1:  # Nº pedido
                ws2[f"{col}{i}"].font = _font(bold=True, size=10, color=COR_VERMELHO)
                ws2[f"{col}{i}"].alignment = _align("center")
            elif j == 2:  # Data
                ws2[f"{col}{i}"].alignment = _align("center")
            elif j == 8:  # Valor
                ws2[f"{col}{i}"].number_format = 'R$ #,##0.00'
                ws2[f"{col}{i}"].font = _font(bold=True, size=10, color=COR_VERDE)
                ws2[f"{col}{i}"].alignment = _align("right")
            elif j in (4,5,6,7,9):
                ws2[f"{col}{i}"].alignment = _align("center")
        ws2.row_dimensions[i].height = 16

    # AutoFilter na aba Pedidos
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(colunas_ped))}{len(pedidos)+1}"

    # Linha de total
    tot_row = len(pedidos) + 2
    ws2[f"A{tot_row}"] = "TOTAL GERAL"
    ws2[f"A{tot_row}"].font = _font(bold=True, size=11, color=COR_BRANCO)
    ws2[f"A{tot_row}"].fill = _fill(COR_VERMELHO)
    for c in range(2, 10):
        ws2[f"{get_column_letter(c)}{tot_row}"].fill = _fill(COR_VERMELHO)
    ws2[f"H{tot_row}"] = total_gasto
    ws2[f"H{tot_row}"].number_format = 'R$ #,##0.00'
    ws2[f"H{tot_row}"].font = _font(bold=True, size=11, color=COR_BRANCO)
    ws2[f"H{tot_row}"].alignment = _align("right")
    ws2.row_dimensions[tot_row].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # ABA 3 — ITENS
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Itens")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"

    colunas_it = [
        ("Nº Pedido", 12), ("Data",       12), ("Fornecedor",    28),
        ("Empresa",   14), ("Categoria",  18), ("Descrição",     50),
        ("Qtde",       9), ("Unid.",        8), ("Vlr Unit.",     14),
        ("Vlr Total", 14),
    ]
    for i, (hdr, larg) in enumerate(colunas_it, 1):
        col = get_column_letter(i)
        ws3.column_dimensions[col].width = larg
        ws3[f"{col}1"] = hdr
        ws3[f"{col}1"].font      = _font(bold=True, size=9, color=COR_BRANCO)
        ws3[f"{col}1"].fill      = _fill(COR_HDR_TAB)
        ws3[f"{col}1"].alignment = _align("center")
        ws3[f"{col}1"].border    = _border_thin()
    ws3.row_dimensions[1].height = 20

    linha = 2
    for p in pedidos:
        def _pg2(key, default=""):
            try: v = p[key]; return v if v is not None else default
            except: return default

        num  = str(_pg2("numero",""))
        data = str(_pg2("data_pedido",""))
        forn = str(_pg2("fornecedor_nome","—")).upper()
        emp  = str(_pg2("empresa_faturadora","—"))
        itens = itens_por_pedido.get(num, [])

        if itens:
            for item in itens:
                bg = COR_ALT1 if linha%2==0 else COR_ALT2
                def _ig(key, default=""):
                    try: v = item[key]; return v if v is not None else default
                    except: return default

                try:
                    qtd = float(_ig("quantidade", 0))
                    qtd_fmt = int(qtd) if qtd == int(qtd) else round(qtd, 3)
                except: qtd_fmt = 0

                row_vals = [
                    f"#{num}", data, forn, emp,
                    str(_ig("categoria","")),
                    str(_ig("descricao","")),
                    qtd_fmt,
                    str(_ig("unidade","")),
                    _fmt_val(_ig("valor_unitario",0)),
                    _fmt_val(_ig("valor_total",0)),
                ]
                for j, val in enumerate(row_vals, 1):
                    col = get_column_letter(j)
                    ws3[f"{col}{linha}"] = val
                    ws3[f"{col}{linha}"].fill   = _fill(bg)
                    ws3[f"{col}{linha}"].border = _border_bottom()
                    ws3[f"{col}{linha}"].font   = _font(size=9)
                    if j == 1:
                        ws3[f"{col}{linha}"].font = _font(bold=True, size=9, color=COR_VERMELHO)
                        ws3[f"{col}{linha}"].alignment = _align("center")
                    elif j in (2,3,4):
                        ws3[f"{col}{linha}"].alignment = _align("center")
                    elif j == 6:
                        ws3[f"{col}{linha}"].alignment = _align("left", wrap=True)
                    elif j in (9,10):
                        ws3[f"{col}{linha}"].number_format = 'R$ #,##0.00'
                        ws3[f"{col}{linha}"].alignment = _align("right")
                        if j == 10:
                            ws3[f"{col}{linha}"].font = _font(bold=True, size=9, color=COR_VERDE)
                    elif j == 7:
                        ws3[f"{col}{linha}"].alignment = _align("center")
                    elif j == 8:
                        ws3[f"{col}{linha}"].alignment = _align("center")
                ws3.row_dimensions[linha].height = 15
                linha += 1
        else:
            # Pedido sem itens — linha única com valor total
            bg = COR_ALT1 if linha%2==0 else COR_ALT2
            row_vals = [f"#{num}", data, forn, emp, "—", "(itens não disponíveis)", "—", "—", "—", _fmt_val(_pg2("valor_total",0))]
            for j, val in enumerate(row_vals, 1):
                col = get_column_letter(j)
                ws3[f"{col}{linha}"] = val
                ws3[f"{col}{linha}"].fill = _fill(COR_AMARELO)
                ws3[f"{col}{linha}"].border = _border_bottom()
                ws3[f"{col}{linha}"].font = _font(size=9, italic=True, color=COR_CINZA_MED)
                if j == 1:
                    ws3[f"{col}{linha}"].font = _font(bold=True, size=9, color=COR_VERMELHO)
                    ws3[f"{col}{linha}"].alignment = _align("center")
                elif j == 10:
                    ws3[f"{col}{linha}"].number_format = 'R$ #,##0.00'
                    ws3[f"{col}{linha}"].alignment = _align("right")
            ws3.row_dimensions[linha].height = 15
            linha += 1

    # AutoFilter na aba Itens
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(colunas_it))}{linha}"

    # Total geral itens
    ws3[f"A{linha}"] = "TOTAL GERAL"
    ws3[f"A{linha}"].font = _font(bold=True, size=10, color=COR_BRANCO)
    ws3[f"A{linha}"].fill = _fill(COR_VERMELHO)
    for c in range(2, 11):
        ws3[f"{get_column_letter(c)}{linha}"].fill = _fill(COR_VERMELHO)
    ws3[f"J{linha}"] = total_gasto
    ws3[f"J{linha}"].number_format = 'R$ #,##0.00'
    ws3[f"J{linha}"].font = _font(bold=True, size=10, color=COR_BRANCO)
    ws3[f"J{linha}"].alignment = _align("right")
    ws3.row_dimensions[linha].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # ABA 4 — RESUMO POR CATEGORIA
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Por Categoria")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 14

    # Título
    ws4.merge_cells("A1:D1")
    ws4["A1"] = "RESUMO POR CATEGORIA"
    ws4["A1"].font      = _font(bold=True, size=13, color=COR_BRANCO)
    ws4["A1"].fill      = _fill(COR_VERMELHO)
    ws4["A1"].alignment = _align("center")
    ws4.row_dimensions[1].height = 24

    ws4.merge_cells("A2:D2")
    ws4["A2"] = nome_obra.upper()
    ws4["A2"].font      = _font(bold=True, size=10, color=COR_CINZA_ESC)
    ws4["A2"].fill      = _fill(COR_CINZA_CLR)
    ws4["A2"].alignment = _align("center")
    ws4.row_dimensions[2].height = 18

    # Agrupa itens por categoria
    from collections import defaultdict
    por_cat = defaultdict(lambda: {"qtd_itens":0, "valor":0.0, "descricoes": set()})

    for p in pedidos:
        def _pg3(key, d=""):
            try: v=p[key]; return v if v is not None else d
            except: return d
        num = str(_pg3("numero",""))
        for item in itens_por_pedido.get(num, []):
            def _ig3(key, d=""):
                try: v=item[key]; return v if v is not None else d
                except: return d
            cat_raw = str(_ig3("categoria","") or "").strip().upper()
            cat = cat_raw if cat_raw and cat_raw not in ("NAN","NONE","") else "SEM CATEGORIA"
            por_cat[cat]["qtd_itens"]  += 1
            por_cat[cat]["valor"]      += _fmt_val(_ig3("valor_total", 0))
            por_cat[cat]["descricoes"].add(str(_ig3("descricao",""))[:50])

    # Cabeçalho da tabela
    row_c = 4
    hdrs = [("Categoria","A"),("Qtd. Itens","B"),("Valor Total","C"),("% do Total","D")]
    for hdr, col in hdrs:
        ws4[f"{col}{row_c}"] = hdr
        ws4[f"{col}{row_c}"].font      = _font(bold=True, size=9, color=COR_BRANCO)
        ws4[f"{col}{row_c}"].fill      = _fill(COR_HDR_TAB)
        ws4[f"{col}{row_c}"].alignment = _align("center")
        ws4[f"{col}{row_c}"].border    = _border_thin()
    ws4.row_dimensions[row_c].height = 18
    row_c += 1

    # Dados por categoria (ordenado por valor)
    for i, (cat, info) in enumerate(sorted(por_cat.items(), key=lambda x: -x[1]["valor"])):
        bg = COR_ALT1 if i%2==0 else COR_ALT2
        pct = (info["valor"]/total_gasto*100) if total_gasto > 0 else 0

        ws4[f"A{row_c}"] = cat
        ws4[f"A{row_c}"].font = _font(bold=True, size=10)
        ws4[f"A{row_c}"].fill = _fill(bg)
        ws4[f"A{row_c}"].border = _border_bottom()

        ws4[f"B{row_c}"] = info["qtd_itens"]
        ws4[f"B{row_c}"].font = _font(size=10)
        ws4[f"B{row_c}"].fill = _fill(bg)
        ws4[f"B{row_c}"].alignment = _align("center")
        ws4[f"B{row_c}"].border = _border_bottom()

        ws4[f"C{row_c}"] = info["valor"]
        ws4[f"C{row_c}"].number_format = 'R$ #,##0.00'
        ws4[f"C{row_c}"].font = _font(bold=True, size=10, color=COR_VERDE)
        ws4[f"C{row_c}"].fill = _fill(bg)
        ws4[f"C{row_c}"].alignment = _align("right")
        ws4[f"C{row_c}"].border = _border_bottom()

        ws4[f"D{row_c}"] = round(pct/100, 4)
        ws4[f"D{row_c}"].number_format = '0.0%'
        ws4[f"D{row_c}"].font = _font(size=10, color=COR_CINZA_MED)
        ws4[f"D{row_c}"].fill = _fill(bg)
        ws4[f"D{row_c}"].alignment = _align("center")
        ws4[f"D{row_c}"].border = _border_bottom()
        ws4.row_dimensions[row_c].height = 16
        row_c += 1

    # AutoFilter na aba categoria
    ws4.auto_filter.ref = f"A4:D{row_c-1}"

    # Total
    ws4[f"A{row_c}"] = "TOTAL"
    ws4[f"A{row_c}"].font = _font(bold=True, size=10, color=COR_BRANCO)
    ws4[f"A{row_c}"].fill = _fill(COR_VERMELHO)
    for col in ["B","C","D"]:
        ws4[f"{col}{row_c}"].fill = _fill(COR_VERMELHO)
    ws4[f"B{row_c}"] = sum(v["qtd_itens"] for v in por_cat.values())
    ws4[f"B{row_c}"].font = _font(bold=True, size=10, color=COR_BRANCO)
    ws4[f"B{row_c}"].alignment = _align("center")
    ws4[f"C{row_c}"] = total_gasto
    ws4[f"C{row_c}"].number_format = 'R$ #,##0.00'
    ws4[f"C{row_c}"].font = _font(bold=True, size=10, color=COR_BRANCO)
    ws4[f"C{row_c}"].alignment = _align("right")
    ws4.row_dimensions[row_c].height = 20

    # Reordena abas: Resumo, Por Categoria, Pedidos, Itens
    wb.move_sheet("Por Categoria", offset=-(wb.sheetnames.index("Por Categoria")-1))

    wb.save(caminho)


def _fmt_brl(v):
    try:
        return f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    except:
        return "0,00"
