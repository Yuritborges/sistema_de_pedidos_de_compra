# app/data/locacoes_import.py
# Leitura da planilha LANÇAMENTO → SQLite compartilhado (locacoes.db).
# Importação automática na inicialização quando configurado.

from __future__ import annotations

import os
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from config import (
    LOCACOES_AUTO_IMPORT_SE_VAZIO,
    LOCACOES_AUTO_SYNC_PLANILHA_NOVA,
    LOCACOES_PLANILHA_CANDIDATES,
    LOCACOES_PLANILHA_ENV,
    LOCACOES_PLANILHA_MANUAL,
)

_LAST_SYNC_USER_MESSAGE: str | None = None


def consume_last_sync_message() -> str | None:
    global _LAST_SYNC_USER_MESSAGE
    msg = _LAST_SYNC_USER_MESSAGE
    _LAST_SYNC_USER_MESSAGE = None
    return msg


def clean_str(v):
    txt = str(v or "").strip()
    if txt.lower() in ("none", "nan", "nat"):
        return ""
    return txt


def to_iso_date(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    txt = str(v).strip()
    if not txt:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(txt[:19], fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    try:
        return datetime.strptime(txt[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except Exception:
        return txt[:10]


def parse_periodo(v):
    if v is None or v == "":
        return None
    try:
        if isinstance(v, (int, float)):
            return int(v)
        s = str(v).strip().replace(",", ".")
        return int(float(s))
    except Exception:
        return None


def calcular_derivados_locacao(
    data_pedido_iso: str,
    periodo_dias,
    data_vencimento_iso: str,
    pedido_ok: str,
):
    """Mesma lógica da planilha / widget de locações."""
    po = clean_str(pedido_ok).upper()
    dp_iso = clean_str(data_pedido_iso)
    per = parse_periodo(periodo_dias)
    v_iso = clean_str(data_vencimento_iso)

    venc_date = None
    if dp_iso and per is not None and per >= 0:
        try:
            d0 = datetime.strptime(dp_iso[:10], "%Y-%m-%d").date()
            venc_date = d0 + timedelta(days=per)
        except Exception:
            venc_date = None
    if venc_date is None and v_iso:
        try:
            venc_date = datetime.strptime(v_iso[:10], "%Y-%m-%d").date()
        except Exception:
            venc_date = None

    if po == "OK":
        v_out = venc_date.strftime("%Y-%m-%d") if venc_date else (v_iso[:10] if v_iso else "")
        return v_out, "OK", "OK"

    if venc_date is None:
        return (v_iso[:10] if v_iso else ""), "", ""

    hoje = date.today()
    delta = (venc_date - hoje).days
    dias_txt = str(delta)
    v_out = venc_date.strftime("%Y-%m-%d")
    if delta < 0:
        sit = "VENCIDO"
    elif delta <= 7:
        sit = "ATUALIZAR"
    else:
        sit = "NA OBRA"
    return v_out, dias_txt, sit


def derivados_locacao_linha(row: dict) -> tuple[str, str, str]:
    """Vencimento, dias_a_vencer e situação com base em hoje (não usa cache desatualizado do BD)."""
    return calcular_derivados_locacao(
        row.get("data_pedido") or "",
        row.get("periodo_dias"),
        row.get("data_vencimento") or "",
        row.get("pedido_ok") or "",
    )


def destaque_visual_linha_locacao_db(row: dict) -> str | None:
    """
    Mesmo critério da tabela Locações (linha vermelha / faixa amarela «2 dias»).
    Retorna None, 'vencido' ou 'dois_dias'.
    Sempre recalcula dias/situação a partir das datas (evita alerta zerado com BD antigo).
    """
    if clean_str(row.get("pedido_ok")).upper() == "OK":
        return None
    _venc, dias_txt, sit = derivados_locacao_linha(row)
    sit = (sit or str(row.get("situacao") or "")).strip().upper()
    d = None
    try:
        d = int(str(dias_txt or "").strip())
    except ValueError:
        try:
            d = int(str(row.get("dias_a_vencer") or "").strip())
        except ValueError:
            pass
    if sit == "VENCIDO":
        return "vencido"
    if d is not None and d < 0:
        return "vencido"
    if d is not None and 0 <= d <= 7:
        return "dois_dias"
    if sit == "ATUALIZAR":
        return "dois_dias"
    return None


def resolver_caminho_planilha_locacoes() -> str | None:
    """Primeiro caminho válido: variável de ambiente → manual em config → pasta do projeto."""
    if LOCACOES_PLANILHA_ENV and os.path.isfile(LOCACOES_PLANILHA_ENV):
        return os.path.normpath(LOCACOES_PLANILHA_ENV)
    manual = (LOCACOES_PLANILHA_MANUAL or "").strip()
    if manual and os.path.isfile(manual):
        return os.path.normpath(manual)
    for p in LOCACOES_PLANILHA_CANDIDATES:
        if p and os.path.isfile(p):
            return os.path.normpath(p)
    return None


def import_locacoes_into_connection(conn, path: str, substituir: bool) -> int:
    """
    Importa linhas do Excel para locacoes_registros.
    `conn` já deve estar aberto (mesmo banco compartilhado).
    Retorna quantidade de linhas inseridas.
    """
    path = os.path.normpath(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        nome_aba = next((s for s in wb.sheetnames if "LAN" in s.upper()), None)
        if not nome_aba:
            raise ValueError("Aba LANÇAMENTO não encontrada.")
        ws = wb[nome_aba]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
    finally:
        wb.close()

    if substituir:
        conn.execute("DELETE FROM locacoes_registros")

    inseridos = 0
    base = os.path.basename(path)
    for row in rows:
        if not row:
            continue
        obra = clean_str(row[0] if len(row) > 0 else "")
        comprador = clean_str(row[1] if len(row) > 1 else "")
        numero_pedido = clean_str(row[2] if len(row) > 2 else "")
        fornecedor = clean_str(row[3] if len(row) > 3 else "")
        item_locado = clean_str(row[4] if len(row) > 4 else "")
        data_pedido = to_iso_date(row[5] if len(row) > 5 else "")
        periodo = parse_periodo(row[6] if len(row) > 6 else None)
        data_venc_plan = to_iso_date(row[7] if len(row) > 7 else "")
        pedido_ok_raw = clean_str(row[10] if len(row) > 10 else "")
        pedido_ok = "OK" if pedido_ok_raw.upper() == "OK" else ""

        if not (obra or numero_pedido or item_locado):
            continue

        venc, dias, sit = calcular_derivados_locacao(
            data_pedido,
            periodo if periodo is not None else "",
            data_venc_plan,
            pedido_ok,
        )

        conn.execute(
            """
            INSERT INTO locacoes_registros (
                obra, comprador, numero_pedido, fornecedor, item_locado, tipo,
                pedido_compra_numero,
                data_pedido, periodo_dias, data_vencimento, dias_a_vencer,
                situacao, pedido_ok, origem_planilha, atualizado_em
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """,
            (
                obra,
                comprador,
                numero_pedido,
                fornecedor,
                item_locado,
                "",
                "",
                data_pedido,
                periodo,
                venc,
                dias,
                sit,
                pedido_ok,
                base,
            ),
        )
        inseridos += 1

    return inseridos


def contar_locacoes_vencimento_e_alerta() -> tuple[int, int]:
    """
    Conta linhas com o mesmo destaque visual da tabela (vencido vs. alerta amarelo).
    Usa colunas gravadas no BD (`situacao`, `dias_a_vencer`, `pedido_ok`).
    Retorna (n_vencidos, n_proximo_vencer).
    """
    try:
        from app.data.database import get_locacoes_connection
    except Exception:
        return 0, 0
    try:
        with get_locacoes_connection() as conn:
            rows = conn.execute(
                """
                SELECT pedido_ok, situacao, dias_a_vencer,
                       data_pedido, periodo_dias, data_vencimento
                FROM locacoes_registros
                """
            ).fetchall()
    except Exception:
        return 0, 0
    venc = 0
    alert = 0
    for r in rows:
        tag = destaque_visual_linha_locacao_db(dict(r))
        if tag == "vencido":
            venc += 1
        elif tag == "dois_dias":
            alert += 1
    return venc, alert


def _meta_get(conn, chave: str) -> str | None:
    row = conn.execute(
        "SELECT valor FROM locacoes_meta WHERE chave=?",
        (chave,),
    ).fetchone()
    if not row:
        return None
    return row[0]


def _meta_set(conn, chave: str, valor: str):
    conn.execute(
        "INSERT OR REPLACE INTO locacoes_meta (chave, valor) VALUES (?, ?)",
        (chave, valor),
    )


def registrar_planilha_na_meta(conn, path: str) -> None:
    """Após importação manual ou automática, grava mtime para uso opcional da sincronização."""
    path = os.path.normpath(path)
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return
    _meta_set(conn, "planilha_mtime_ultima_importacao", str(mtime))
    _meta_set(conn, "planilha_caminho_ultima_importacao", path)


def tentar_sincronizar_planilha_locacoes_no_startup() -> None:
    """
    Chamado após init_locacoes_shared_db():
    - Se o BD está vazio e LOCACOES_AUTO_IMPORT_SE_VAZIO: importa a planilha padrão.
    - Se LOCACOES_AUTO_SYNC_PLANILHA_NOVA e o arquivo no disco é mais recente que a
      última importação gravada em meta: substitui todo o conteúdo (uso consciente).
    """
    global _LAST_SYNC_USER_MESSAGE

    path = resolver_caminho_planilha_locacoes()
    if not path:
        return

    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return

    from app.data.database import get_locacoes_connection

    with get_locacoes_connection() as conn:
        row_cnt = conn.execute(
            "SELECT COUNT(*) FROM locacoes_registros",
        ).fetchone()
        count = int(row_cnt[0]) if row_cnt else 0

        meta_mtime_s = _meta_get(conn, "planilha_mtime_ultima_importacao") or ""
        try:
            meta_mtime = float(meta_mtime_s)
        except ValueError:
            meta_mtime = 0.0

        substituir = False
        motivo = ""

        if count == 0 and LOCACOES_AUTO_IMPORT_SE_VAZIO:
            substituir = True
            motivo = "banco vazio"
        elif (
            LOCACOES_AUTO_SYNC_PLANILHA_NOVA
            and count > 0
            and meta_mtime > 0
            and mtime > meta_mtime
        ):
            substituir = True
            motivo = "planilha atualizada no disco"

        if not substituir:
            return

        inseridos = import_locacoes_into_connection(conn, path, substituir=True)
        registrar_planilha_na_meta(conn, path)

    _LAST_SYNC_USER_MESSAGE = (
        f"Importação automática ({motivo}): {inseridos} registros da planilha → banco."
    )
    try:
        print(f"[Locações] {_LAST_SYNC_USER_MESSAGE}")
    except Exception:
        pass
